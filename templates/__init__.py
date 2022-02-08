# -*- coding: utf-8 -*-
"""
Contains base template class for report objects.

Example
-------
import os

from reportio import ReportTemplate


# Create report object at runtime
class Report(ReportTemplate):

    def __init__(self,
                 report_name: str = 'test',
                 log_location: str = os.path.join(
                     os.path.dirname(__file__), 'simple_log.txt'),
                 config_location: str = os.path.join(
                     os.path.dirname(__file__), 'simple_config.txt'),
                 connection_dictionary: Dict[str, object] = {},
                 client: dd.Client = None,
                 optional_function: callable = None) -> None:
        super().__init__(report_name,
                         log_location,
                         config_location,
                         connection_dictionary,
                         client,
                         optional_function)

    # 'run' method must be instantiated
    def run(self):
        write_config()
        self.file = self.test_get_data(connection=self.test_get_connection())[0]
        self.test_backup_data()


# Script report object to run
if __name__ == '__main__':
    report = Report()
    try:
        report.run()
    except Exception:
        report.log("See log for debug details", 'CRITICAL')
        report.backup_data()
        report.log("Backup successful")
        input("PRESS ENTER KEY TO QUIT")
        report.log("QUITTING")
"""


# %% Imports
# %%% Py3 Standard
from abc import ABC, abstractmethod
import sys
import os
from getpass import getpass
import shutil
import multiprocessing
import configparser as cfg
from datetime import date as dt_date
from tempfile import NamedTemporaryFile
import sqlite3
from typing import List, Dict, Any, Tuple

# %%% 3rd Party
import pandas as pd
# TODO: Remove this import when updating get_connection function
import pyodbc
from dask import delayed as dask_delayed
import dask.distributed as dd
from openpyxl import load_workbook
# Pending tqdm.dask module release
# from tqdm.dask import TqdmCallback as ProgressBar

# %%% User-Defined
from reportio import logger
from reportio.data import Data
from reportio.errors import (ConfigError,
                             ReportNameError,
                             DBConnectionError,
                             DatasetNameError,
                             UnexpectedDbType)
from reportio.future import ProgressBar


# %% Variables
# %%% Private
__all__ = ['write_config', 'ReportTemplate']

# %%% Public
default_self_location: str = os.path.dirname(__file__)
default_config_location: str = os.path.join(
    default_self_location, 'config.txt')
default_log_location: str = os.path.join(
    os.path.dirname(sys.argv[0]), 'log.txt')
# TODO: find a way to initialize client here without causing:
#     TypeError: can't pickle _asyncio.Task objects in data module
# client = dd.Client(processes=False)
client = None


# %% Classes
class ReportTemplate(ABC):
    """
    Abstract template class for building custom reports.

    Developer must implement at least the 'run' method.

    Parameters
    ----------
    report_name : string
        Name of report.
    log_location : directory, default is in user folder
        File where script will log processes. Will create new log file if
        directory does not exist.
    config_location : directory, default is template config file
        Location of config file. If location does not point to a file, a
        copy of the default config will be created in this location.
    connection_dictionary : dictionary, optional
        Active connection objects for report to use if desired.
    client : dask.distributed.Client
        From dask.distributed module. Used to schedule multiprocessing.
    optional_function : function, optional
        Will run just before attempting resume.
    """

    # %%% Functions
    # %%%% Private
    def __init__(self,
                 report_name: str,
                 log_location: str = default_log_location,
                 config_location: str = default_config_location,
                 connection_dictionary: Dict[str, object] = {},
                 client: dd.Client = client,
                 optional_function: callable = None,
                 _default_config_location: str = default_config_location
                 ) -> None:
        # Assign base variables
        self.report_name: str = report_name
        self.log_location: str = log_location
        self.config_location: str = config_location
        self.connection_dictionary: Dict[str, object] = connection_dictionary
        self.client: dd.Client = client
        self.start_date: str = dt_date.today().isoformat()
        self._sheets: int = 1
        self._files: List[str] = []
        # Configure logger
        new_log: str = logger.config(logger.getLogger(__name__),
                                     self.log_location)
        self.log: callable = logger.log
        # Check if report name can be used
        try:
            NamedTemporaryFile(suffix='__' + self.report_name).close()
        except OSError:
            raise ReportNameError(self.report_name)
        # Write config
        config_tuple: tuple = write_config(self.config_location,
                                           self.report_name)
        self.config: cfg.ConfigParser = config_tuple[0]
        self.self_location: str = config_tuple[1]
        config_created: bool = config_tuple[2]
        del config_tuple
        if config_created:
            self.log("Creating config file at: '{0}'".format(config_location))
        # Make backup dir if needed
        self.backup_folder_location: str =\
            self.config['REPORT']['backup_folder']
        if not os.path.isdir(self.backup_folder_location):
            self.log("Creating backup dir at '{0}'".format(
                self.backup_folder_location), 'DEBUG')
            os.makedirs(self.backup_folder_location)
        # Make temporary files dir if needed
        self.temp_files_location: str = \
            self.config['REPORT']['temp_files_folder']
        if not os.path.isdir(self.temp_files_location):
            self.log("Creating temp dir at '{0}'".format(
                self.temp_files_location))
            os.makedirs(self.temp_files_location)
        # Inform user
        self.log("Starting report with {0} log located at '{1}'"
                 .format(new_log, log_location))
        self.log("""Variables:
                     dirSelf: '{0}'
                     log_location: '{1}'
                     config_location: '{2}'""".format(
                     self.self_location, log_location, config_location),
                 'DEBUG')
        # Run optional function if callable
        if callable(optional_function):
            optional_function()
        # Delete old report if it exists
        report_location: str = self.config['REPORT']['export_to']
        if os.path.exists(report_location):
            self.log("Removing {0}".format(report_location))
            os.remove(report_location)
        else:
            report_folder: str = os.path.dirname(report_location)
            file_list: List[str] = [os.path.join(
                report_folder, file_name) for file_name in os.listdir(
                    report_folder) if os.path.isfile(
                        os.path.join(report_folder, file_name))]
            for file_name in file_list:
                if report_location.split('.')[0] in file_name:
                    self.log("Removing {0}".format(file_name))
                    os.remove(file_name)
        # Attempt resume
        self.files: List[str] = self.attempt_resume()

    def _delete_old_file(self, file_location: str) -> None:
        if os.path.exists(file_location):
            os.remove(file_location)

    # %%%% Public
    def backup_data(self, optional_function: callable = None) -> None:
        """
        Save temporary files to backup folder in the event of CRICITAL error.

        Backup files saved in this was are utilized if script is run again the
        same day as the backup.

        Parameters
        ----------
        optional_function : function, optional
            Will be run after files are copied but before this function is
            completed.
        """
        self.log("Backing up data", 'WARNING')
        self.log("Backup location: {0}".format(self.backup_folder_location),
                 'DEBUG')
        # Copy parquet files
        for file in self._files:
            if hasattr(file, 'name'):
                if file.name.split('.')[-1] == 'gz' and os.path.isfile(
                        file.name):
                    try:
                        destination_location: str = os.path.join(
                            self.backup_folder_location,
                            file.name.split('__')[-1])
                        self.log("Backing up '{0}' to '{1}'".format(
                            file.name, file.name), 'DEBUG')
                        pd.read_parquet(file).to_parquet(destination_location,
                                                         compression='gzip')
                    # If one backup fails, try other files
                    except AttributeError:
                        continue
        # Run optional function if included
        if callable(optional_function):
            optional_function()
        # Write text file with start date
        with open(os.path.join(self.backup_folder_location, 'startDate.txt'),
                  'w') as file:
            file.write(self.start_date)

    def delete_data_backup(self, optional_function: callable = None) -> None:
        """
        Delete all gz files in backup directory.

        Parameters
        ----------
        optional_function : function, optional
            Will be run for each file in backup dir with file directory as a
            parameter.
        """
        # Loop files in backup directory
        self.log("Cleaning up old data backup")
        for file_location in [os.path.join(self.backup_folder_location, s) for
                              s in os.listdir(self.backup_folder_location)]:
            # Run optional function if included
            if callable(optional_function):
                optional_function(file_location)
            if file_location.split('.')[-1] in ['gz', 'txt']:
                self.log("Removing '{0}'".format(file_location), 'DEBUG')
                os.remove(file_location)

    def attempt_resume(self,
                       optional_function_1: callable = None,
                       optional_function_2: callable = None) -> List[str]:
        """
        Load files from backup folder.

        Attempts to read files and resume report near the point of error. Only
        loads files that have been backed up today after a CRITICAL error.

        Parameters
        ----------
        optional_function_1 : function, optional
            Will be run if data backup is found.
        optional_function_2 : function, optional
            Will be run if data backup is not from today as optional_function
            parameter in self.delete_data_backup.

        Returns
        -------
        file_list : list
            Files found in backup.
        """
        self.log("Checking for backup files", 'DEBUG')
        file_list: List[str] = []
        # Look for date stamp file
        file_name: str = 'startDate.txt'
        if file_name in os.listdir(self.backup_folder_location):
            with open(os.path.join(
                    self.backup_folder_location, file_name), 'r') as objFile:
                start_date: str = objFile.read()
            # Check if date stamp matches today
            if start_date == dt_date.today().isoformat():
                self.log("Resuming previous attempt")
                # Gather names of files to read from backup
                for file_name in [f for f in os.listdir(
                        self.backup_folder_location
                        ) if f.split('.')[-1] == 'gz']:
                    file_list.append(file_name)
                self.log("These files will be read from backup: {0}".format(
                    file_list), 'DEBUG')
                # Run optional function if included
                if callable(optional_function_1):
                    optional_function_1()
            else:
                self.log("No recent backup files found", 'DEBUG')
                if callable(optional_function_2):
                    self.delete_data_backup(optional_function_2)
                else:
                    self.delete_data_backup()
        else:
            self.log("No backup found", 'DEBUG')
        return file_list

    def get_connection(self,
                       database_name: str,
                       connection_string: str,
                       connection_type: str = 'odbc') -> object:
        """
        Create connection object.

        Will use appropriate connection library and string from config file
        dependant upon conneciton type. Will prompt user for UID/PWD as needed.

        Parameters
        ----------
        database_name : string
            Database name.
        connection_string : string
            Connection string. Must be formatted for database.
        connection_type : ('odbc', 'jdbc', 'sqlite'), Optional
            Specifies the connection type.

        Returns
        -------
        connection_object : connection object
        """
        # Initialize login iteration
        i = 0
        # Loop until connected
        while database_name not in self.connection_dictionary.keys():
            try:
                self.log("Connecting to {0}".format(database_name))
                if database_name == 'sqlite':
                    connection_object = sqlite3.connect(
                        connection_string, check_same_thread=False)
                else:
                    connection_object = pyodbc.connect(connection_string)
                self.connection_dictionary[database_name] = connection_object
                self.log("Connection successful", 'DEBUG')
            # If login fails, get user id/password and retry
            except pyodbc.Error as err:
                i += 1
                # Strip previous user entry if present
                connection_string = 'DSN' + connection_string.split('DSN')[-1]
                # Cap login attempts at 5
                if i > 5:
                    del user_id
                    del password
                    self.log(err, 'DEBUG')
                    self.log("Connection string (without UID/PWD): {0}"
                             .format(connection_string), 'DEBUG')
                    raise DBConnectionError
                self.log("Unable to connect!", 'ERROR')
                self.log("Attempting login with UID/PWD from user.", 'DEBUG')
                user_id: str = input("Enter USERNAME: ")
                password: str = getpass("Enter PASSWORD: ")
                connection_string =  connection_string + ';UID=' + user_id +\
                    ';PWD=' + password
        connection_object = self.connection_dictionary.get(database_name)
        self.log("Using connection object '{0}'".format(connection_object),
                 'DEBUG')
        return connection_object

    # TODO: test
    def _get_connection(self,
                        database: str,
                        connection_type: str = 'sqlite') -> object:
        """
        Create connection object.

        Will use appropriate connection library and string from config file
        dependant upon conneciton type. Will prompt user for UID/PWD as needed.

        Parameters
        ----------
        database_name : string
            Database name as used in config file.
        connection_type : ('odbc', 'jdbc', 'sqlite'), default 'sqlite'
            Specifies the connection type.

        Returns
        -------
        connection_object : connection object
        """
        # Initialize login iteration
        i = 0
        # Loop until connected
        while database not in self.connection_dictionary.keys():
            self.log("Connecting to {0}".format(database))
            if connection_type == 'sqlite':
                connection_string = \
                    self.config[database]['db_location']
                connection_object = sqlite3.connect(
                    connection_string, check_same_thread=False)
            elif connection_type == 'odbc':
                import pyodbc
                # Prevent DB pooling, which causes errors
                if pyodbc.pooling:
                    pyodbc.pooling = False
                connection_string = \
                    self.config[database]['connection_string']
                try:
                    connection_object = pyodbc.connect(connection_string)
                except pyodbc.Error as err:
                    i += 1
                    # Strip previous user entry if present
                    connection_string = 'DSN' +\
                        connection_string.split('DSN')[-1]
                    # Cap login attempts at 5
                    if i > 5:
                        del user_id
                        del password
                        self.log(err, 'DEBUG')
                        self.log("Connection string (without UID/PWD): {0}"
                                 .format(connection_string), 'DEBUG')
                        raise DBConnectionError
                    self.log("Unable to connect!", 'ERROR')
                    self.log(
                        "Attempting login with UID/PWD from user.", 'DEBUG')
                    user_id: str = input("Enter USERNAME: ")
                    password: str = getpass("Enter PASSWORD: ")
                    connection_string = connection_string + ';UID=' + \
                        user_id + ';PWD=' + password
            elif connection_type == 'jdbc':
                import jaydebeapi as jdbc
                if self.config[database]['password'] == '':
                    _password: str = getpass("Enter PASSWORD: ")
                else:
                    _password: str = self.config[database]['password']
                connection_object = jdbc.connect(
                    self.config[database]['class_name'],
                    self.config[database]['url'],
                    {'user': self.config[database]['user'],
                     'password': _password},
                    self.config[database]['jars'])
                del _password
            self.connection_dictionary[database] = connection_object
            self.log("Connection successful", 'DEBUG')
        connection_object = self.connection_dictionary.get(database)
        self.log("Using connection object '{0}'".format(connection_object),
                 'DEBUG')
        return connection_object

    def get_writer(self, report_location: str) -> pd.ExcelWriter:
        """
        Create ExcelWriter object to allow multiple tabs to be written.

        Parameters
        ----------
        report_location : str
            Location of result report.

        Returns
        -------
        excel_writer : pandas.ExcelWriter
            From pandas module utilizing openpyxl as engine.
        """
        # Creates blank Excel file if needed
        if not os.path.isfile(report_location):
            pd.DataFrame().to_excel(report_location)
        # Creates Excel Writer
        excel_writer: pd.ExcelWriter = pd.ExcelWriter(
            report_location, engine='openpyxl', mode='a')
        excel_writer.book = load_workbook(report_location)
        excel_writer.sheets = dict((ws.title, ws) for ws in
                                   excel_writer.book.worksheets)
        return excel_writer

    def get_temp_file(self,
                      file_name: str,
                      data: pd.DataFrame,
                      folder_location: str = None) -> object:
        """
        Create tempfile object and write dataframe to it.

        File will always use gzip compression and end in .gz. Raises
        DatasetNameError if file_name cannot be used in file. Overwriting may
        cause a critical error and data corruption.

        Parameters
        ----------
        file_name : string
            Name appended to end of temporary file.
        data : DataFrame
            Data to be written to temporary file.
        folder_location : directory, default is location in config
            Folder where file is stored.

        Returns
        -------
        file : object
            Contains data from data. From tempfile module.
        """
        if folder_location is None:
            folder_location = self.temp_files_location
        try:
            file: object = NamedTemporaryFile(dir=folder_location,
                                              suffix="__" + file_name + '.gz')
        except OSError as err:
            self.log(err, 'DEBUG')
            raise DatasetNameError(file_name)
        data.to_parquet(file, compression='gzip')
        self._files.append(file)
        return file

    # TODO: create a 'data' object for reporting module to reduce ambiguity
    def get_data(self,
                 data_name: str,
                 sql: str,
                 db_type: str = '',
                 connection_object: object = None,
                 database_location: str = '') -> Tuple[object, object]:
        """
        Retrieve data from database via connection_object.

        Parameters
        ----------
        data_name : string
            Name for dataset. Cannot use '__', for file path.
        sql : string
            SQL statement as a string. Should be formatted for desired database
        db_type : [sections in config file], optional
            Either db_type or connection_object must be provided to connect to
            database.
        connection_object : connection_object, optional
            If not provided, report will attempt to connect using string in
            config vile. Either db_type or connection_object must be provided
            to connect to database. From pyodbc module.
        database_location : string, optional
             Database location. Must only be provided if db_type is
             'access'.

        Returns
        -------
        data : (file, pandas.DataFrame)
            Contains data from query.
        connection_object : object
            Database connection_object.
        """
        # Check config file for connection_object string
        if db_type not in self.config['DB']:
            self.log("Config location: {0}".format(self.config_location),
                     'DEBUG')
            raise UnexpectedDbType(db_type)
        backup_file_location: str = os.path.join(self.backup_folder_location,
                                                 data_name + '.gz')
        if not os.path.isfile(backup_file_location):
            # Create new connection_object if needed
            if connection_object is None:
                connection_object = self.get_connection(
                    db_type, self.config['DB'][db_type])
            self.log("Querying database")
            temporary_dataframe: pd.DataFrame = pd.read_sql(
                sql, connection_object)
        else:
            self.log("Reading backup file")
            temporary_dataframe: pd.DataFrame = pd.read_parquet(
                backup_file_location)
        if len(temporary_dataframe.index) == 0:
            self.log("Query was empty", 'WARNING')
        if data_name != '':
            data: object = self.get_temp_file(data_name, temporary_dataframe)
        else:
            data: object = temporary_dataframe
        return data, connection_object

    def _get_data(self,
                  dataset_name: str,
                  sql: str = '',
                  connection: object = None) -> Data:
        return Data(dataset_name,
                    self.log,
                    self.client,
                    self.temp_files_location,
                    self.backup_folder_location,
                    sql,
                    connection)

    def cross_query(self,
                    dataframe_input: pd.DataFrame,
                    column_list: List[str],
                    sql_function: callable,
                    db_type: str = '',
                    connection_object: object = None,
                    db_location: str = '',
                    zero_value: Any = pd.NA,
                    final_columns: List[str] = None,
                    compress_columns: List[str] = None) -> object:
        """
        For experimental use.

        Append column_list to dataframe_input from data from another query.
        Wil dynamically build queries based on values in each row of
        dataframe_input and execute while utilizing dask for multithread
        scheduling.

        Parameters
        ----------
        dataframe_input : pandas.DataFrame
            DataFrame to be appended with data from queries.
        column_list : list
            Name of new columns to be added.
        sql_function : function
            Should have input for row of data from DataFrame.iterrows() and
            return query as string.
        db_type : {{0}}, optional
            Either db_type or connection_object must be provided to connect to
            database.
        connection_object : object, optional
            If not provided, report will attempt to connect using string in
            config vile. Either db_type or connection_object must be provided
            to connect to database.
        db_location : str, optional
             Database location. Must only be provided if db_type is
             'access'.
        zero_value : user-defined type, default is pandas.NA
            Value to be used if query returns no results.
        final_columns : list, optional
            Columns listed in result file.
        compress_columns : list, optional
            Column names to use as a temporary grouping. One query will be
            run for each unique record in this column list. This is untested.

        Returns
        -------
        dataframe_ouput : DataFrame
            dataframe_input with data populated in column_list from
            sql_function.
        """
        def process_chunk(dataframe_input: pd.DataFrame,
                          chunk_list: List[tuple],
                          column_list: List[Any] = column_list,
                          connection_object: object = connection_object,
                          compress_columns: List[Any] = compress_columns
                          ) -> pd.DataFrame:
            # Initialize chunk dataframe
            chunk: pd.DataFrame = pd.DataFrame(columns=column_list)
            # Loop chunk list
            for index, row in chunk_list:
                # Check for partial population
                if not any([pd.isna(dataframe_input.at[index, c])
                            for c in column_list]):
                    chunk.append(
                        pd.DataFrame([list(dataframe_input.loc[index])],
                                     index=[index],
                                     columns=chunk.columns))
                else:
                    data_tuple: Tuple[pd.DataFrame, object] = \
                        self.get_data(
                            '', sql_function(row), db_type, connection_object)
                    query_data: pd.DataFrame = data_tuple[0]
                    connection_object: object = data_tuple[1]
                    del data_tuple
                    if len(query_data.index) > 0:
                        for i in query_data.index:
                            chunk = chunk.append(
                                pd.DataFrame([list(query_data.loc[i])],
                                             index=[index],
                                             columns=chunk.columns))
                    else:
                        chunk.append(
                            pd.DataFrame(
                                [[zero_value] * len(column_list)],
                                index=[index],
                                columns=chunk.columns))
            if compress_columns is not None:
                # Initialize output DataFrame
                ungrouped_data: pd.DataFrame = pd.DataFrame(
                    columns=chunk.columns)
                # Duplicate current index
                chunk.reset_index()
                # Loop over input
                for _, r in chunk.iterrows():
                    # Iterate over grouped data
                    for i in range(len(r['index'])):
                        ungrouped_data = ungrouped_data.append(
                            pd.DataFrame([r[column_list]],
                                         index=[r['index'][i]],
                                         columns=column_list))
                        for c in compress_columns:
                            ungrouped_data.at[r['index'][i], c] = r[c][i]
                chunk = ungrouped_data
            return chunk
        # Add new columns if needed
        if any([c not in l for l in dataframe_input.columns for c in
                column_list]):
            for c in column_list:
                if c not in dataframe_input.columns:
                    dataframe_input[c] = pd.NA
        # Temporarily group data if needed (this may not work)
        if compress_columns is not None:
            dataframe_input = dataframe_input.groupby(
                compress_columns, as_index=False).agg(list)
        # Initialize chunk list, dataframe list
        chunk_list: List[tuple] = []
        dataframe_list: List[pd.DataFrame] = []
        # Loop dataframe
        for index_row in dataframe_input.iterrows():
            chunk_list.append(index_row)
            # Process in chunks equal to number of threads available
            if len(chunk_list) >= len(
                    dataframe_input.index) / multiprocessing.cpu_count():
                dataframe_list.append(
                    dask_delayed(process_chunk)(dataframe_input, chunk_list))
                chunk_list = []
        # Attempt connection before spamming with all threads
        self.get_connection(db_type, self.config['DB'][db_type])
        with ProgressBar():
            dask_delayed(len)(dataframe_list).compute()
        # Initialize output dataframe
        dataframe_ouput = dataframe_input.copy()
        for df in dataframe_list:
            # Combine master df and chunk df, keeping chunk where not null
            dataframe_ouput.combine(df, lambda s1, s2: s1.combine(
                s2, lambda v1, v2: v2 if not pd.isna(v2) else v1),
                overwrite=False)
        if final_columns is not None:
            dataframe_ouput = dataframe_ouput.loc[:, final_columns]
        return dataframe_ouput

    def merge_files(self,
                    file_name: str,
                    file_1: object,
                    file_2: object,
                    merge_column: str,
                    join_type: str = 'inner',
                    columns_1: List[str] = None,
                    columns_2: List[str] = None,
                    columns_final: List[str] = None) -> object:
        """
        Merge two parquet files into one.

        Parameters
        ----------
        file_name : string
            Output file name.
        file_1 : file
            Contains data from pandas.
        file_2 : file
            Contains data from pandas.
        merge_column : string
            Column used for merge.
        join_type : {'left', 'right', 'outer', 'inner'}, default 'inner'
            Merge type.
        columns_1 : list, default all columns
            Columns merged from first file.
        columns_2 : list, default all columns
            Columns merged from second file.
        columns_final : list, default all columns
            Columns listed in result file.

        Returns
        -------
        data : object, merged file
        """
        # Check for backup file
        backup_file_location: str = os.path.join(self.backup_folder_location,
                                                 file_name + '.gz')
        if os.path.isfile(backup_file_location):
            self.log("Reading backup file")
            temporary_dataframe: pd.DataFrame = pd.read_parquet(
                backup_file_location)
            if file_name != '':
                data: object = self.get_temp_file(file_name,
                                                  temporary_dataframe)
        else:
            self.log("Merging files")
            dataframe_1: pd.DataFrame = pd.read_parquet(file_1,
                                                        columns=columns_1)
            dataframe_2: pd.DataFrame = pd.read_parquet(file_2,
                                                        columns=columns_2)
            if join_type == 'right':
                suffix_right: str = ''
                suffix_left: str = "_drop"
            else:
                suffix_right: str = '_drop'
                suffix_left: str = ''
            dataframe_1 = dataframe_1.merge(
                dataframe_2, join_type, merge_column, suffixes=(
                    suffix_left, suffix_right))
            del dataframe_2
            dataframe_1 = dataframe_1.drop(
                [c for c in dataframe_1 if '_drop' in c], axis=1)
            if columns_final is not None:
                dataframe_1 = dataframe_1[columns_final]
            dataframe_1 = dataframe_1.drop_duplicates(ignore_index=True)
            data: object = self.get_temp_file(file_name, dataframe_1)
        return data

    def export_data(self,
                    file: object,
                    report_location: str,
                    sheet: str = '',
                    excel_writer: pd.ExcelWriter = None) -> str:
        """
        Export report data to excel file.

        Will change file type to CSV as needed.

        Parameters
        ----------
        file : object
            Parquet file. Report data for output.
        report_location : str
            Location of result file. File extension will be overwritten
        sheet : str, optional
            Name of excel sheet. Will append to file name if data too large for
            Excel.
        excel_writer : pd.ExcelWriter, optional
            For use if writing multiple tabs. From pandas module.

        Returns
        -------
        report_location : str
            Final directory used for export.
        """
        # Check file format
        if len(report_location.split('.')) > 1:
            self.log("File extension will be overwritten", 'WARNING')
            self.log(
                "Attempted report location: '{0}'".format(report_location),
                'DEBUG')
            report_location = report_location.split('.')[0]
        # Read data file
        data: pd.DataFrame = pd.read_parquet(file)
        # Check file size
        if len(data.index) > 1048576 or len(data.columns) > 16384:
            self.log("Exporting data to CSV")
            report_location = report_location + "__" + sheet + '.csv'
            data.to_csv(report_location, index=False)
        else:
            self.log("Exporting data to Excel")
            # Handle user variables
            if sheet == '':
                sheet = 'Sheet' + str(self._sheets)
            report_location = report_location + '.xlsx'
            # Write to Excel
            if excel_writer is None:
                self.excel_writer: pd.ExcelWriter = self.get_writer(
                    report_location)
            else:
                self.excel_writer: pd.ExcelWriter = excel_writer
            data.to_excel(self.excel_writer, sheet, index=False)
            self.excel_writer.save()
            self.excel_writer.close()
        self._sheets += 1
        return report_location

    @abstractmethod
    def run(self):
        """Must implement this method to run report."""
        pass


class ReportTemplate_2(ABC):
    """
    Abstract template class for building custom reports.

    Developer must implement at least the 'run' method.

    Parameters
    ----------
    report_name : string
        Name of report.
    log_location : directory, default is in user folder
        File where script will log processes. Will create new log file if
        directory does not exist.
    config_location : directory, default is template config file
        Location of config file. If location does not point to a file, a
        copy of the default config will be created in this location.
    connection_dictionary : dictionary, optional
        Active connection objects for report to use if desired.
    client : dask.distributed.Client
        From dask.distributed module. Used to schedule multiprocessing.
    optional_function : function, optional
        Will run just before attempting resume.
    """

    # %%% Functions
    # %%%% Private
    def __init__(self,
                 report_name: str,
                 log_location: str = default_log_location,
                 config_location: str = default_config_location,
                 connection_dictionary: Dict[str, object] = {},
                 client: dd.Client = client,
                 optional_function: callable = None
                 ) -> None:
        # Assign base variables
        self.report_name: str = report_name
        self.log_location: str = log_location
        self.config_location: str = config_location
        self.connection_dictionary: Dict[str, object] = connection_dictionary
        self.client: dd.Client = client
        self.start_date: str = dt_date.today().isoformat()
        self._sheets: int = 1
        self._files: List[str] = []
        # Configure logger
        new_log: str = logger.config(logger.getLogger(__name__),
                                     self.log_location)
        self.log: callable = logger.log
        # Check if report name can be used
        try:
            NamedTemporaryFile(suffix='__' + self.report_name).close()
        except OSError:
            raise ReportNameError(self.report_name)
        # Write config
        config_tuple: tuple = write_config(self.config_location,
                                           self.report_name)
        self.config: cfg.ConfigParser = config_tuple[0]
        self.self_location: str = config_tuple[1]
        config_created: bool = config_tuple[2]
        del config_tuple
        if config_created:
            self.log("Creating config file at: '{0}'".format(config_location))
        # Make backup dir if needed
        self.backup_folder_location: str =\
            self.config['REPORT']['backup_folder']
        if not os.path.isdir(self.backup_folder_location):
            self.log("Creating backup dir at '{0}'".format(
                self.backup_folder_location), 'DEBUG')
            os.makedirs(self.backup_folder_location)
        # Make temporary files dir if needed
        self.temp_files_location: str = \
            self.config['REPORT']['temp_files_folder']
        if not os.path.isdir(self.temp_files_location):
            self.log("Creating temp dir at '{0}'".format(
                self.temp_files_location))
            os.makedirs(self.temp_files_location)
        # Inform user
        self.log("Starting report with {0} log located at '{1}'"
                 .format(new_log, log_location))
        self.log("""Variables:
                     dirSelf: '{0}'
                     log_location: '{1}'
                     config_location: '{2}'""".format(
                     self.self_location, log_location, config_location),
                 'DEBUG')
        # Run optional function if callable
        if callable(optional_function):
            optional_function()
        # Delete old report if it exists
        report_location: str = self.config['REPORT']['export_to']
        if os.path.exists(report_location):
            self.log("Removing {0}".format(report_location))
            os.remove(report_location)
        else:
            report_folder: str = os.path.dirname(report_location)
            file_list: List[str] = [os.path.join(
                report_folder, file_name) for file_name in os.listdir(
                    report_folder) if os.path.isfile(
                        os.path.join(report_folder, file_name))]
            for file_name in file_list:
                if report_location.split('.')[0] in file_name:
                    self.log("Removing {0}".format(file_name))
                    os.remove(file_name)
        # Attempt resume
        self.files: List[str] = self.attempt_resume()

    def _delete_old_file(self, file_location: str) -> None:
        if os.path.exists(file_location):
            os.remove(file_location)

    # %%%% Public
    def backup_data(self, optional_function: callable = None) -> None:
        """
        Save temporary files to backup folder in the event of CRICITAL error.

        Backup files saved in this was are utilized if script is run again the
        same day as the backup.

        Parameters
        ----------
        optional_function : function, optional
            Will be run after files are copied but before this function is
            completed.
        """
        self.log("Backing up data", 'WARNING')
        self.log("Backup location: {0}".format(self.backup_folder_location),
                 'DEBUG')
        # Copy parquet files
        for file in self._files:
            if hasattr(file, 'name'):
                if file.name.split('.')[-1] == 'gz' and os.path.isfile(
                        file.name):
                    try:
                        destination_location: str = os.path.join(
                            self.backup_folder_location,
                            file.name.split('__')[-1])
                        self.log("Backing up '{0}' to '{1}'".format(
                            file.name, file.name), 'DEBUG')
                        pd.read_parquet(file).to_parquet(destination_location,
                                                         compression='gzip')
                    # If one backup fails, try other files
                    except AttributeError:
                        continue
        # Run optional function if included
        if callable(optional_function):
            optional_function()
        # Write text file with start date
        with open(os.path.join(self.backup_folder_location, 'startDate.txt'),
                  'w') as file:
            file.write(self.start_date)

    def delete_data_backup(self, optional_function: callable = None) -> None:
        """
        Delete all gz files in backup directory.

        Parameters
        ----------
        optional_function : function, optional
            Will be run for each file in backup dir with file directory as a
            parameter.
        """
        # Loop files in backup directory
        self.log("Cleaning up old data backup")
        for file_location in [os.path.join(self.backup_folder_location, s) for
                              s in os.listdir(self.backup_folder_location)]:
            # Run optional function if included
            if callable(optional_function):
                optional_function(file_location)
            if file_location.split('.')[-1] in ['gz', 'txt']:
                self.log("Removing '{0}'".format(file_location), 'DEBUG')
                os.remove(file_location)

    def attempt_resume(self,
                       optional_function_1: callable = None,
                       optional_function_2: callable = None) -> List[str]:
        """
        Load files from backup folder.

        Attempts to read files and resume report near the point of error. Only
        loads files that have been backed up today after a CRITICAL error.

        Parameters
        ----------
        optional_function_1 : function, optional
            Will be run if data backup is found.
        optional_function_2 : function, optional
            Will be run if data backup is not from today as optional_function
            parameter in self.delete_data_backup.

        Returns
        -------
        file_list : list
            Files found in backup.
        """
        self.log("Checking for backup files", 'DEBUG')
        file_list: List[str] = []
        # Look for date stamp file
        file_name: str = 'startDate.txt'
        if file_name in os.listdir(self.backup_folder_location):
            with open(os.path.join(
                    self.backup_folder_location, file_name), 'r') as objFile:
                start_date: str = objFile.read()
            # Check if date stamp matches today
            if start_date == dt_date.today().isoformat():
                self.log("Resuming previous attempt")
                # Gather names of files to read from backup
                for file_name in [f for f in os.listdir(
                        self.backup_folder_location
                        ) if f.split('.')[-1] == 'gz']:
                    file_list.append(file_name)
                self.log("These files will be read from backup: {0}".format(
                    file_list), 'DEBUG')
                # Run optional function if included
                if callable(optional_function_1):
                    optional_function_1()
            else:
                self.log("No recent backup files found", 'DEBUG')
                if callable(optional_function_2):
                    self.delete_data_backup(optional_function_2)
                else:
                    self.delete_data_backup()
        else:
            self.log("No backup found", 'DEBUG')
        return file_list

    def get_connection(self,
                       database_name: str,
                       connection_string: str,
                       connection_type: str = 'odbc') -> object:
        """
        Create connection object.

        Will use appropriate connection library and string from config file
        dependant upon conneciton type. Will prompt user for UID/PWD as needed.

        Parameters
        ----------
        database_name : string
            Database name.
        connection_string : string
            Connection string. Must be formatted for database.
        connection_type : ('odbc', 'jdbc', 'sqlite'), Optional
            Specifies the connection type.

        Returns
        -------
        connection_object : connection object
        """
        # Initialize login iteration
        i = 0
        # Loop until connected
        while database_name not in self.connection_dictionary.keys():
            try:
                self.log("Connecting to {0}".format(database_name))
                if database_name == 'sqlite':
                    connection_object = sqlite3.connect(
                        connection_string, check_same_thread=False)
                else:
                    connection_object = pyodbc.connect(connection_string)
                self.connection_dictionary[database_name] = connection_object
                self.log("Connection successful", 'DEBUG')
            # If login fails, get user id/password and retry
            except pyodbc.Error as err:
                i += 1
                # Strip previous user entry if present
                connection_string = 'DSN' + connection_string.split('DSN')[-1]
                # Cap login attempts at 5
                if i > 5:
                    del user_id
                    del password
                    self.log(err, 'DEBUG')
                    self.log("Connection string (without UID/PWD): {0}"
                             .format(connection_string), 'DEBUG')
                    raise DBConnectionError
                self.log("Unable to connect!", 'ERROR')
                self.log("Attempting login with UID/PWD from user.", 'DEBUG')
                user_id: str = input("Enter USERNAME: ")
                password: str = input("Enter PASSWORD: ")
                connection_string = 'UID=' + user_id +\
                    ';PWD=' + password + ';' + connection_string
        connection_object = self.connection_dictionary.get(database_name)
        self.log("Using connection object '{0}'".format(connection_object),
                 'DEBUG')
        return connection_object

    # TODO: test
    def _get_connection(self,
                        database: str,
                        connection_type: str = 'sqlite') -> object:
        """
        Create connection object.

        Will use appropriate connection library and string from config file
        dependant upon conneciton type. Will prompt user for UID/PWD as needed.

        Parameters
        ----------
        database_name : string
            Database name as used in config file.
        connection_type : ('odbc', 'jdbc', 'sqlite'), default 'sqlite'
            Specifies the connection type.

        Returns
        -------
        connection_object : connection object
        """
        # Initialize login iteration
        i = 0
        # Loop until connected
        while database not in self.connection_dictionary.keys():
            self.log("Connecting to {0}".format(database))
            if connection_type == 'sqlite':
                connection_string = \
                    self.config[database]['db_location']
                connection_object = sqlite3.connect(
                    connection_string, check_same_thread=False)
            elif connection_type == 'odbc':
                import pyodbc
                # Prevent DB pooling, which causes errors
                if pyodbc.pooling:
                    pyodbc.pooling = False
                connection_string = \
                    self.config[database]['connection_string']
                try:
                    connection_object = pyodbc.connect(connection_string)
                except pyodbc.Error as err:
                    i += 1
                    # Strip previous user entry if present
                    connection_string = 'DSN' +\
                        connection_string.split('DSN')[-1]
                    # Cap login attempts at 5
                    if i > 5:
                        del user_id
                        del password
                        self.log(err, 'DEBUG')
                        self.log("Connection string (without UID/PWD): {0}"
                                 .format(connection_string), 'DEBUG')
                        raise DBConnectionError
                    self.log("Unable to connect!", 'ERROR')
                    self.log(
                        "Attempting login with UID/PWD from user.", 'DEBUG')
                    user_id: str = input("Enter USERNAME: ")
                    password: str = getpass("Enter PASSWORD: ")
                    connection_string = connection_string + ';UID=' + \
                        user_id + ';PWD=' + password
            elif connection_type == 'jdbc':
                import jaydebeapi as jdbc
                if self.config[database]['password'] == '':
                    _password: str = getpass("Enter PASSWORD: ")
                else:
                    _password: str = self.config[database]['password']
                connection_object = jdbc.connect(
                    self.config[database]['class_name'],
                    self.config[database]['url'],
                    {'user': self.config[database]['user'],
                     'password': _password},
                    self.config[database]['jars'])
                del _password
            self.connection_dictionary[database] = connection_object
            self.log("Connection successful", 'DEBUG')
        connection_object = self.connection_dictionary.get(database)
        self.log("Using connection object '{0}'".format(connection_object),
                 'DEBUG')
        return connection_object

    def get_writer(self, report_location: str) -> pd.ExcelWriter:
        """
        Create ExcelWriter object to allow multiple tabs to be written.

        Parameters
        ----------
        report_location : str
            Location of result report.

        Returns
        -------
        excel_writer : pandas.ExcelWriter
            From pandas module utilizing openpyxl as engine.
        """
        # Creates blank Excel file if needed
        if not os.path.isfile(report_location):
            pd.DataFrame().to_excel(report_location)
        # Creates Excel Writer
        excel_writer: pd.ExcelWriter = pd.ExcelWriter(
            report_location, engine='openpyxl', mode='a')
        excel_writer.book = load_workbook(report_location)
        excel_writer.sheets = dict((ws.title, ws) for ws in
                                   excel_writer.book.worksheets)
        return excel_writer

    def get_temp_file(self,
                      file_name: str,
                      data: pd.DataFrame,
                      folder_location: str = None) -> object:
        """
        Create tempfile object and write dataframe to it.

        File will always use gzip compression and end in .gz. Raises
        DatasetNameError if file_name cannot be used in file. Overwriting may
        cause a critical error and data corruption.

        Parameters
        ----------
        file_name : string
            Name appended to end of temporary file.
        data : DataFrame
            Data to be written to temporary file.
        folder_location : directory, default is location in config
            Folder where file is stored.

        Returns
        -------
        file : object
            Contains data from data. From tempfile module.
        """
        if folder_location is None:
            folder_location = self.temp_files_location
        try:
            file: object = NamedTemporaryFile(dir=folder_location,
                                              suffix="__" + file_name + '.gz')
        except OSError as err:
            self.log(err, 'DEBUG')
            raise DatasetNameError(file_name)
        data.to_parquet(file, compression='gzip')
        self._files.append(file)
        return file

    def export_data(self,
                    file: object,
                    report_location: str,
                    sheet: str = '',
                    excel_writer: pd.ExcelWriter = None) -> str:
        """
        Export report data to excel file.

        Will change file type to CSV as needed.

        Parameters
        ----------
        file : object
            Parquet file. Report data for output.
        report_location : str
            Location of result file. File extension will be overwritten
        sheet : str, optional
            Name of excel sheet. Will append to file name if data too large for
            Excel.
        excel_writer : pd.ExcelWriter, optional
            For use if writing multiple tabs. From pandas module.

        Returns
        -------
        report_location : str
            Final directory used for export.
        """
        # Check file format
        if len(report_location.split('.')) > 1:
            self.log("File extension will be overwritten", 'WARNING')
            self.log(
                "Attempted report location: '{0}'".format(report_location),
                'DEBUG')
            report_location = report_location.split('.')[0]
        # Read data file
        data: pd.DataFrame = pd.read_parquet(file)
        # Check file size
        if len(data.index) > 1048576 or len(data.columns) > 16384:
            self.log("Exporting data to CSV")
            report_location = report_location + "__" + sheet + '.csv'
            data.to_csv(report_location, index=False)
        else:
            self.log("Exporting data to Excel")
            # Handle user variables
            if sheet == '':
                sheet = 'Sheet' + str(self._sheets)
            report_location = report_location + '.xlsx'
            # Write to Excel
            if excel_writer is None:
                self.excel_writer: pd.ExcelWriter = self.get_writer(
                    report_location)
            else:
                self.excel_writer: pd.ExcelWriter = excel_writer
            data.to_excel(self.excel_writer, sheet, index=False)
            self.excel_writer.save()
            self.excel_writer.close()
        self._sheets += 1
        return report_location

    @abstractmethod
    def run(self):
        """Must implement this method to run report."""
        pass


# %% Functions
def write_config(config_location: str = default_config_location,
                 report_name: str = 'reserved for use at runtime'):
    """
    Rewrites config file with current location and report name.

    Parameters
    ----------
    config_location : str, optional
        DESCRIPTION. The default is default_config_location.
    report_name : str, optional
        DESCRIPTION. The default is 'reserved for use at runtime'.

    Returns
    -------
    config : configparser.ConfigParser
        Parsed config file.
    self_location : str
        Location of config file.
    config_created : bool
        True if new config file was created.
    """
    try:
        # Read config file if present
        config: cfg.ConfigParser = cfg.ConfigParser()
        config._interpolation = cfg.ExtendedInterpolation()
        config_created: bool = False
        if not os.path.isfile(config_location):
            config_created = True
            global default_config_location
            shutil.copy(default_config_location, config_location)
        config.read(config_location)
        # Write unpopulated variables
        if sys.argv[0] == '':
            global default_self_location
            self_location = default_self_location
        else:
            self_location = os.path.dirname(sys.argv[0])
        config['DEFAULT']['self_dir'] = default_self_location
        config['DEFAULT']['self_folder'] = os.path.dirname(
            default_self_location)
        config['PATHS']['self_dir'] = self_location
        config['PATHS']['self_folder'] = os.path.dirname(self_location)
        config['REPORT']['report_name'] = report_name
        with open(config_location, 'w') as file:
            config.write(file)
    except Exception:
        raise ConfigError
    return config, self_location, config_created


# %% Script
write_config()
